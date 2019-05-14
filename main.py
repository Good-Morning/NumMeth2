import os, sys, numpy, random, copy, solver, re
import openpyxl, itertools

def goodCond(size):
    return [[10/(1 + 4*abs(i - j)) for j in range(size)] for i in range(size)]

def randCond(size):
    return [[100*random.random() for i in range(size)] for j in range(size)]

def gilbCond(size):
    return [[1/(2+i+j) for i in range(size)] for j in range(size)]

def normCond(size):
    t = randCond(size)
    n = numpy.linalg.norm(t)
    return [[x/n for x in xs] for xs in t]

modules = {}

for module in map(lambda x: x[:-3], filter(lambda x: x != 'main.py' and x[-3:] == '.py', os.listdir())):
    exec('import {}'.format(module))
    mod = sys.modules[module]
    modules[module] = [x for x in [getattr(mod, x) for x in dir(mod)] if str(type(x)) == "<class 'type'>" and issubclass(x, solver.Solver)]
modules = {i:modules[i] for i in modules if len(modules[i])}

sizeFrom = 5
sizeTo = 5

tests = [
    ('random',    [randCond(i) for i in range(sizeFrom, sizeTo+1)]), 
    ('good cond', [goodCond(i) for i in range(sizeFrom, sizeTo+1)]), 
    ('normized', [normCond(i) for i in range(sizeFrom, sizeTo+1)]), 
    ('bad cond',  [gilbCond(i) for i in range(sizeFrom, sizeTo+1)]), 
]

wb = openpyxl.Workbook()
ws = wb['Sheet']

sizeSize = sizeTo - sizeFrom + 1
side = openpyxl.styles.Side(border_style='thin', color="FF000000")
zide = openpyxl.styles.Side(border_style='medium', color="FF000000")
ws['A1'] = 'module'
ws['B1'] = 'solver'
ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
ws['B1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
ws['A1'].border = openpyxl.styles.Border(right=side, bottom=zide)
ws['B1'].border = openpyxl.styles.Border(right=zide, bottom=zide)
ws.merge_cells('A1:A3')
ws.merge_cells('B1:B3')
head = 3
height = 3
for module in modules:
    height += len(modules[module])
for test in tests:
    ws.cell(1, head, test[0])
    ws.cell(1, head).alignment = openpyxl.styles.Alignment(horizontal='center')
    ws.cell(1, head).border = openpyxl.styles.Border(right=zide)
    for i in range(3, height + 1):
        ws.cell(i, head).border = openpyxl.styles.Border(right=side)
        ws.cell(i, head+1).border = openpyxl.styles.Border(right=side)
    ws.merge_cells(start_column=head, start_row=1, end_column=head + sum(range(sizeFrom+2, sizeTo+3)) - 1, end_row=1)
    for i in range(sizeFrom, sizeTo + 1):
        ws.cell(2, head, 'size: ' + str(i) + '   cond: ' + '{0:.6g}'.format(numpy.linalg.cond(test[1][i - sizeFrom])))
        ws.cell(2, head).alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.cell(2, head).border = openpyxl.styles.Border(right=zide)
        ws.cell(3, head,   'error')
        ws.cell(3, head+1, 'iterations')
        ws.cell(3, head+2, 'solution')
        ws.cell(3, head).border   = openpyxl.styles.Border(left=side, right=side, bottom=zide)
        ws.cell(3, head+1).border = openpyxl.styles.Border(left=side, right=side, bottom=zide)
        ws.cell(3, head+2).border = openpyxl.styles.Border(left=side, right=zide, bottom=zide)
        ws.merge_cells(None, 3, head+2, 3, head + i + 1)
        ws.merge_cells(start_column=head, start_row=2, end_column=head + i + 1, end_row=2)
        for j in range(3, height + 1):
            was = ws.cell(j, head + i + 1).border
            ws.cell(j, head + i + 1).border = openpyxl.styles.Border(right=zide, top=was.top, bottom=was.bottom)
        head += i+2
methods = 0
width = head
for module in modules:
    ws.cell(4+methods, 1, module)
    ws.cell(4+methods, 1).alignment = openpyxl.styles.Alignment(vertical='top')
    ws.cell(4+methods, 1).border = openpyxl.styles.Border(right=side, bottom=side)
    bottom = 3+methods+len(modules[module])
    for i in range(1, width):
        was = ws.cell(bottom, i).border
        ws.cell(bottom, i).border = openpyxl.styles.Border(bottom=side, right=was.right, left=was.left, top=was.top)
    ws.merge_cells(None, 4+methods, 1, bottom, 1)
    for worker in modules[module]:
        ws.cell(4+methods, 2, re.match(r"<class '.*?\.(.*)'>", str(worker))[1])
        was = ws.cell(4+methods, 2).border
        ws.cell(4+methods, 2).border = openpyxl.styles.Border(right=zide, left=was.left, bottom=was.bottom, top=was.top)
        head = 3
        for testPack in tests:
            for A in testPack[1]:
                b = [i for i in range(len(A))]
                fun = worker().solve
                (x, iters) = fun(copy.deepcopy(A), copy.deepcopy(b))
                ws.cell(4+methods, head, '{0:.4g}'.format(numpy.linalg.norm(b - numpy.dot(A, x))))
                ws.cell(4+methods, head + 1, str(iters))
                head += 2
                for i in x:
                    ws.cell(4+methods, head, str(i))
                    head += 1
        methods += 1

for cc in ws.columns:
    ws.column_dimensions[cc[3].column_letter].width = max(2 + len(c.value or "") for c in cc)

wb.save('output.xlsx')
